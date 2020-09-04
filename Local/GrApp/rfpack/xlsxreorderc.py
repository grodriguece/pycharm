from pathlib import Path
from openpyxl import load_workbook


def neworder(file, fpos, tpos):
    """Takes a list of ints, and inserts the fpos (from position) int, to tpos (to position)"""
    wb = load_workbook(filename=file)
    shlist = wb.sheetnames  # # Get current order sheets in workbook
    lst = []
    lpos = (len(shlist) - 1)
    print("Before:", [x for x in range(len(shlist))])
    if lpos >= fpos > tpos >= 0:  # move from a high to low position
        for x in range(lpos+1):
            if x == tpos:
                lst.append(fpos)
            elif tpos < x <= fpos:
                lst.append(x-1)
            else:
                lst.append(x)
    if lpos >= tpos > fpos >= 0:  # move from a low to high position
        for x in range(lpos+1):
            if x == tpos:
                lst.append(fpos)
            elif fpos <= x < tpos:
                lst.append(x+1)
            else:
                lst.append(x)
    print("After:", [x for x in range(len(lst))])
    wb._sheets = [wb._sheets[i] for i in lst]  # get each object instance from  wb._sheets, and replace
    wb.save(filename=file)
    return


filex = Path('C:/SQLite/Feat2.xlsx')
neworder(filex, 0, 85)
